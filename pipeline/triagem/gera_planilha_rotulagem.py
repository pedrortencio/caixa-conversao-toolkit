"""Gera a planilha xlsx de rotulagem humana da coluna `registro`.

Lê `dados/triagem/amostra_para_rotular.csv`, filtra as peças `status==keep`
e escreve um xlsx com dropdown (incidental/operacional_rotina/substantivo),
cabeçalho congelado e larguras de coluna para leitura. O xlsx evita as
armadilhas de encoding do Excel com CSV (abre e salva UTF-8 sem corromper
acentos). Os rótulos voltam ao CSV canônico depois, por `item_id`.

Uso: uv run python pipeline/triagem/gera_planilha_rotulagem.py
"""

import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ENTRADA = Path("dados/triagem/amostra_para_rotular.csv")
SAIDA = Path("dados/triagem/rotulagem_registro.xlsx")

CATEGORIAS = ["incidental", "operacional_rotina", "substantivo"]

# Ordem de leitura para rotulagem; `registro` por último, ao lado do texto.
COLUNAS = [
    "item_id", "jornal", "data", "data_confiavel", "page_number",
    "forma", "titulo", "secao", "trecho_caixa", "texto", "registro",
]

LARGURAS = {
    "item_id": 26, "jornal": 16, "data": 11, "data_confiavel": 6,
    "page_number": 6, "forma": 14, "titulo": 28, "secao": 16,
    "trecho_caixa": 45, "texto": 90, "registro": 20,
}

# openpyxl rejeita caracteres de controle (sobras de OCR ruim); remover.
_CONTROLE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _limpa(valor: str) -> str:
    return _CONTROLE.sub("", valor)


def gera_planilha(entrada: Path, saida: Path) -> int:
    """Escreve o xlsx só com as linhas keep; retorna quantas escreveu."""
    with open(entrada, encoding="utf-8", newline="") as f:
        linhas = [r for r in csv.DictReader(f) if r["status"] == "keep"]

    wb = Workbook()
    ws = wb.active
    ws.title = "rotulagem"

    ws.append(COLUNAS)
    for celula in ws[1]:
        celula.font = Font(bold=True)

    for r in linhas:
        # `registro` sai vazio: o rótulo é julgamento humano do Pedro.
        ws.append([_limpa(r[c]) if c != "registro" else None for c in COLUNAS])

    quebra = Alignment(wrap_text=True, vertical="top")
    for idx, nome in enumerate(COLUNAS, start=1):
        letra = get_column_letter(idx)
        ws.column_dimensions[letra].width = LARGURAS[nome]
        if nome in ("trecho_caixa", "texto", "titulo"):
            for celula in ws[letra][1:]:
                celula.alignment = quebra

    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(CATEGORIAS) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valor inválido",
        error="Use uma das três categorias de registro.",
    )
    col_registro = get_column_letter(COLUNAS.index("registro") + 1)
    dv.add(f"{col_registro}2:{col_registro}{len(linhas) + 1}")
    ws.add_data_validation(dv)

    ws.freeze_panes = "A2"
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    return len(linhas)


def main() -> int:
    n = gera_planilha(ENTRADA, SAIDA)
    print(f"{n} peças keep escritas em {SAIDA}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
