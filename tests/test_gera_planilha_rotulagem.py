"""Testes do gerador da planilha de rotulagem de registro (xlsx)."""

import csv

from openpyxl import load_workbook

from pipeline.triagem.gera_planilha_rotulagem import COLUNAS, gera_planilha

CABECALHO = [
    "item_id", "jornal", "source_identifier", "source_year", "page_number",
    "data", "data_fonte", "titulo", "secao", "forma", "continua",
    "trecho_caixa", "texto", "match_ids_pagina", "ocr_contexto",
    "modelo", "protocolo", "registro", "status", "data_confiavel",
]


def _linha(item_id, status, titulo="Título çãé", texto="corpo"):
    base = {c: "" for c in CABECALHO}
    base.update({
        "item_id": item_id, "jornal": "o_paiz", "source_year": "1906",
        "page_number": "1", "data": "1906-01-13", "titulo": titulo,
        "forma": "artigo", "trecho_caixa": "Caixa de Conversão",
        "texto": texto, "status": status, "data_confiavel": "1",
    })
    return base


def _escreve_csv(caminho, linhas):
    with open(caminho, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CABECALHO)
        w.writeheader()
        w.writerows(linhas)


def test_gera_planilha_so_keep_com_dropdown(tmp_path):
    entrada = tmp_path / "amostra.csv"
    saida = tmp_path / "rotulagem.xlsx"
    _escreve_csv(entrada, [
        _linha("a:p001:i0", "keep"),
        _linha("b:p001:i0", "drop_vizinho"),
        _linha("c:p002:i1", "keep", texto="com nul \x00 e controle \x08 no meio"),
    ])

    n = gera_planilha(entrada, saida)

    assert n == 2
    wb = load_workbook(saida)
    ws = wb.active
    # cabeçalho na ordem definida + só as linhas keep
    assert [c.value for c in ws[1]] == COLUNAS
    assert ws.max_row == 3
    ids = {ws.cell(row=r, column=1).value for r in (2, 3)}
    assert ids == {"a:p001:i0", "c:p002:i1"}
    # acentos preservados
    col_titulo = COLUNAS.index("titulo") + 1
    assert ws.cell(row=2, column=col_titulo).value == "Título çãé"
    # caracteres de controle removidos (openpyxl os rejeita)
    col_texto = COLUNAS.index("texto") + 1
    assert ws.cell(row=3, column=col_texto).value == "com nul  e controle  no meio"
    # registro vazio com dropdown das três categorias
    col_registro = COLUNAS.index("registro") + 1
    assert ws.cell(row=2, column=col_registro).value is None
    dvs = list(ws.data_validations.dataValidation)
    assert len(dvs) == 1
    assert dvs[0].formula1 == '"incidental,operacional_rotina,substantivo"'
